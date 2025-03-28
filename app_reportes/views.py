# Django imports
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, F, Q
from django.utils import timezone

# Third-party imports
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

# Local imports
from app_usuarios.utils import is_admin_or_superuser
from app_inventario.models import Producto
from app_ventas.models import Venta, VentaDetalle
from app_finanzas.models import Ingreso, Egreso
from app_pedidos.models import Pedido

# Helper functions
def estilizar_encabezado(celda):
    """Aplica estilos al encabezado de Excel."""
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def productos_mas_vendidos(tipo_tiempo='mensual'):
    """Retorna los productos más vendidos en un periodo."""
    hoy = timezone.now().date()
    inicio_periodo = (
        hoy - timezone.timedelta(days=7)
        if tipo_tiempo == 'semanal'
        else hoy.replace(day=1)
    )

    return VentaDetalle.objects.filter(
        venta__fecha__range=[inicio_periodo, hoy]
    ).values('producto__nombre').annotate(
        total_vendido=Sum('cantidad')
    ).order_by('-total_vendido')

def productos_sin_stock():
    """Retorna productos sin stock."""
    return Producto.objects.filter(cantidad_stock=0)

def productos_no_vendidos():
    """Retorna productos sin ventas en los últimos 30 días."""
    hace_30_dias = timezone.now().date() - timezone.timedelta(days=30)
    return Producto.objects.exclude(
        id__in=VentaDetalle.objects.filter(
            venta__fecha__gte=hace_30_dias
        ).values('producto')
    )

# View functions
@login_required
@user_passes_test(is_admin_or_superuser)
def reporte_inventario(request):
    context = {
        'productos_mas_vendidos': VentaDetalle.objects.values(
            'producto__nombre'
        ).annotate(
            total_vendido=Sum('cantidad')
        ).order_by('-total_vendido')[:10],
        'productos_sin_stock': productos_sin_stock(),
        'productos_agotandose': Producto.objects.filter(
            cantidad_stock__lt=F('stock_minimo')
        ).order_by('cantidad_stock'),
        'productos_no_vendidos': Producto.objects.exclude(
            ventadetalle__isnull=False
        ),
    }
    return render(request, 'reportes/reporte_inventario.html', context)

@login_required
@user_passes_test(is_admin_or_superuser)
def reporte_ingresos_egresos(request):
    tipo_tiempo = request.GET.get('tipo_tiempo', 'mensual')
    hoy = timezone.now().date()
    
    inicio_periodo = (
        hoy - timezone.timedelta(days=7)
        if tipo_tiempo == 'semanal'
        else hoy.replace(day=1)
    )
    
    inicio_periodo = timezone.make_aware(
        timezone.datetime.combine(inicio_periodo, timezone.datetime.min.time())
    )
    fin_periodo = timezone.make_aware(
        timezone.datetime.combine(hoy, timezone.datetime.max.time())
    )

    ingresos = Venta.objects.filter(
        fecha__range=[inicio_periodo, fin_periodo]
    ).aggregate(
        total_ingresos=Sum('total')
    )['total_ingresos'] or 0

    egresos = Egreso.objects.filter(
        fecha__range=[inicio_periodo, fin_periodo]
    ).aggregate(
        total_egresos=Sum('monto')
    )['total_egresos'] or 0

    context = {
        'ingresos': ingresos,
        'egresos': egresos,
        'balance': ingresos - egresos,
        'tipo_tiempo': tipo_tiempo,
    }
    return render(request, 'reportes/reporte_ingresos_egresos.html', context)

@login_required
@user_passes_test(is_admin_or_superuser)
def exportar_reporte_excel(request):
    hoy = timezone.now().date()

    # Creamos un nuevo libro de Excel
    wb = openpyxl.Workbook()

    # --- Hoja 1: Productos más vendidos ---
    ws1 = wb.active
    ws1.title = 'Productos más vendidos'

    # Encabezados
    ws1.append(['Producto', 'Total Vendido'])
    for celda in ws1[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_mas_vendidos = VentaDetalle.objects.values('producto__nombre').annotate(
        total_vendido=Sum('cantidad')).order_by('-total_vendido')[:10]
    for producto in productos_mas_vendidos:
        ws1.append([producto['producto__nombre'], producto['total_vendido']])

    # Ajustar ancho de columnas
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Gráfico de barras para productos más vendidos
    chart1 = BarChart()
    data = Reference(ws1, min_col=2, min_row=1, max_row=len(productos_mas_vendidos) + 1)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=len(productos_mas_vendidos) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Productos más vendidos"
    chart1.y_axis.title = "Cantidad vendida"
    chart1.x_axis.title = "Producto"
    ws1.add_chart(chart1, "E5")

    # --- Hoja 2: Productos sin stock ---
    ws2 = wb.create_sheet('Productos sin stock')

    # Encabezados
    ws2.append(['Producto'])
    for celda in ws2[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_sin_stock = Producto.objects.filter(cantidad_stock=0)
    for producto in productos_sin_stock:
        ws2.append([producto.nombre])

    # Ajustar ancho de columnas
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Hoja 3: Productos que se agotan más rápido ---
    ws3 = wb.create_sheet('Productos agotándose')

    # Encabezados
    ws3.append(['Producto', 'Stock actual'])
    for celda in ws3[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_agotandose = Producto.objects.filter(cantidad_stock__lt=F('stock_minimo')).order_by('cantidad_stock')
    for producto in productos_agotandose:
        ws3.append([producto.nombre, producto.cantidad_stock])

    # Gráfico de barras para productos que se agotan más rápido
    chart2 = BarChart()
    data = Reference(ws3, min_col=2, min_row=1, max_row=len(productos_agotandose) + 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(productos_agotandose) + 1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.title = "Productos con bajo stock"
    chart2.y_axis.title = "Stock actual"
    chart2.x_axis.title = "Producto"
    ws3.add_chart(chart2, "E5")

    # --- Hoja 4: Productos que no se venden ---
    ws4 = wb.create_sheet('Productos no vendidos')

    # Encabezados
    ws4.append(['Producto'])
    for celda in ws4[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_no_vendidos = Producto.objects.exclude(ventadetalle__isnull=False)
    for producto in productos_no_vendidos:
        ws4.append([producto.nombre])


    # Configurar la respuesta HTTP para la descarga de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'

    # Guardar el archivo en el response
    wb.save(response)
    return response

def home(request):
    context = {
        'productos_bajo_stock_count': Producto.objects.filter(
            cantidad_stock__lt=F('stock_minimo')
        ).count(),
        'productos_sin_stock_count': productos_sin_stock().count(),
        'pedidos_pendientes_count': Pedido.objects.filter(
            Q(estado='pedido') | Q(estado='en camino')
        ).count(),
    }
    return render(request, 'home.html', context)

